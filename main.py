from PyQt6.QtCore import *
from PyQt6.QtWidgets import *
import openpyxl
import sys
import time

# Изменение метода __lt__ чтобы можно было производить поиск текстом
class TableWidgetItem(QTableWidgetItem):
    def __lt__(self, other):
        try:
            return float(self.text()) < float(other.text())
        except ValueError:
            return super().__lt__(other)

PROGRAM_TEXT = "Lorem ipsum dolor sit amet, consectetur adipiscing elit. " \
       "Nullam malesuada tellus in ex elementum, vitae rutrum enim vestibulum."

AUTHOR_TEXT = "Lorem ipsum dolor sit amet, consectetur adipiscing elit. " \
       "Nullam malesuada tellus in ex elementum, vitae rutrum enim vestibulum."

class InfoWindow(QWidget):
    def __init__(self):
        super().__init__()

        self.TextLabel = QLabel()
        self.TextLabel.setWordWrap(True)
        self.TextLabel.setMinimumSize(200, 200)
        self._layount = QVBoxLayout()

        self._layount.addWidget(self.TextLabel)

        self.setLayout(self._layount)

    def ShowText(self, text):
        self.show()
        self.TextLabel.setText(text)


# Основное окно приложения
class Window(QMainWindow):
    def __init__(self):
        super().__init__()

        self.menu_bar = QMenuBar()

        self.InfoWindow = InfoWindow()

        self.menu = QMenu("Файл")
        self.info = QMenu("Инфо")
        action1 = self.menu.addAction("Сохранить")
        action2 = self.menu.addAction("Сохранить лист")

        author = self.info.addAction("Об Авторе")
        programm = self.info.addAction("Об программе")

        action1.triggered.connect(self.export)
        action2.triggered.connect(self.exportSheet)
        author.triggered.connect(self.authorInfo)
        programm.triggered.connect(self.programmInfo)

        self.menu_bar.addMenu(self.menu)
        self.menu_bar.addMenu(self.info)

        self.setMenuBar(self.menu_bar)

    def authorInfo(self):
        self.InfoWindow.ShowText(AUTHOR_TEXT)

    def programmInfo(self):
        self.InfoWindow.ShowText(PROGRAM_TEXT)

    def closeEvent(self, event):
        widget.DataWindow.close()

        event.accept()

    def exportSheet(self):
        filename, filter = QFileDialog.getSaveFileName(self, 'Save file', '','Excel files (*.xlsx)')
        wb = openpyxl.Workbook()

        temp = wb["Sheet"]
        wb.remove(temp)

        print(filename)

        centralWidget = self.centralWidget()

        table_widget = widget.currentWidget
        workingSheet = wb.create_sheet("Export")

        labels = []
        for c in range(table_widget.columnCount()):
            it = table_widget.horizontalHeaderItem(c)
            labels.append(str(c+1) if it is None else it.text())
        print(labels)

        for i in range(len(labels)):
            value = labels[i]
            workingSheet.cell(1, i + 1, value)

        if filename:
            for column in range(table_widget.columnCount()):
                k = 1
                for row in range(table_widget.rowCount()):
                    print(row, table_widget.isRowHidden(row))
                    if table_widget.isRowHidden(row) == False:
                        try:
                            text = str(table_widget.item(row, column).text())
                            k += 1
                            workingSheet.cell(k, column + 1, text)
                        except AttributeError:
                            pass
                        
        wb.save(filename)

    def export(self):
        filename, filter = QFileDialog.getSaveFileName(self, 'Save file', '','Excel files (*.xlsx)')
        wb = openpyxl.Workbook()

        temp = wb["Sheet"]
        wb.remove(temp)

        print(filename)

        centralWidget = self.centralWidget()

        for sheet, table_widget in centralWidget.temp_sheets.items():
            workingSheet = wb.create_sheet(sheet.title)

            labels = []
            for c in range(table_widget.columnCount()):
                it = table_widget.horizontalHeaderItem(c)
                labels.append(str(c+1) if it is None else it.text())
            print(labels)

            for i in range(len(labels)):
                value = labels[i]
                workingSheet.cell(1, i + 1, value)

            if filename:
                for column in range(table_widget.columnCount()):
                    for row in range(table_widget.rowCount()):
                        try:
                            text = str(table_widget.item(row, column).text())
                            workingSheet.cell(row + 2, column + 1, text)
                        except AttributeError:
                            pass
                        
        wb.save(filename)

class AcceptRemoveLineWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Подтверждение")
        self.setFixedSize(300, 100)

        self.label = QLabel("Вы точно хотите удалить?")
        self.yes_button = QPushButton("Да")
        self.no_button = QPushButton("Нет")

        self._layout = QVBoxLayout()
        self.setLayout(self._layout)

        self._layout.addWidget(self.label)
        self._layout.addWidget(self.yes_button)
        self._layout.addWidget(self.no_button)

        self.yes_button.clicked.connect(self.accept)
        self.no_button.clicked.connect(self.decline)

    def accept(self):
        widget.currentWidget.removeRow(widget.selectedRow)
        self.close()
    
    def decline(self):
        self.close()

class AcceptWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Подтверждение")
        self.setFixedSize(300, 100)

        self.label = QLabel("Вы точно хотите добавить?")
        self.yes_button = QPushButton("Да")
        self.no_button = QPushButton("Нет")

        self._layout = QVBoxLayout()
        self.setLayout(self._layout)

        self._layout.addWidget(self.label)
        self._layout.addWidget(self.yes_button)
        self._layout.addWidget(self.no_button)

        self.yes_button.clicked.connect(self.accept)
        self.no_button.clicked.connect(self.decline)

    def accept(self):
        widget.currentWidget.insertRow(0)

        for i in range(len(widget.DataWindow.labels)):
            print(0, i)
            item = TableWidgetItem(str(widget.DataWindow.inputs[i].text()))

            widget.currentWidget.setItem(0 , i, item)

        widget.DataWindow.close()
        self.close()
    
    def decline(self):
        self.close()

# Окно для добавления новой даты
class DataWindow(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Add Data")
        self.setFixedSize(500, 500)

        self.Layout = QVBoxLayout()
        self.setLayout(self.Layout)

        self.acceptWindow = AcceptWindow()

        self.addButton = QPushButton("Добавить")
        self.addButton.clicked.connect(self.acceptInfo)
        self.Layout.addWidget(self.addButton)
        
        self.inputs :list[QLineEdit] = []
        self.labels :list[QLabel] = []

    # Добавления новой строки по записанным данным
    def acceptInfo(self):

        self.acceptWindow.show()

    #Загрузка полей для заполнений
    def load(self):
        print("loading stuff")

        self.Layout.removeWidget(self.addButton)

        labels = []

        for c in range(widget.currentWidget.columnCount()):
            label = widget.currentWidget.horizontalHeaderItem(c)
            labels.append(str(c+1) if label is None else label.text())
        
        for l in range(len(labels)):
            label = QLabel(labels[l])
            lineEdit = QLineEdit()
            self.Layout.addWidget(label)
            self.Layout.addWidget(lineEdit)

            self.inputs.append(lineEdit)
            self.labels.append(label)

        self.Layout.addWidget(self.addButton)

    # Действия при закрытии окна
    def closeEvent(self, event):
        print("clearing stuff")

        for i in reversed(range(self.Layout.count())): 
            self.Layout.itemAt(i).widget().setParent(None)

        event.accept()

# Главное окно где создаются списки и редактируется дата
class Main(QWidget):
    def __init__(self):
        super(Main, self).__init__()
        self.setWindowTitle("Load Excel data to QTableWidget")
        self.setBaseSize(700, 700)
        
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        self.addData = QPushButton("Добавить Информацию")
        self.add_row = QPushButton("Добавить линию")
        self.remove_row = QPushButton("Удалить линию")
    
        self.filter_list = QComboBox()

        self.DataWindow = DataWindow()
        self.acceptRemoveWindow = AcceptRemoveLineWindow()

        self.addData.clicked.connect(self.show_data_window)

        self.add_row.clicked.connect(self.addRow)
        self.remove_row.clicked.connect(self.deleteRow)

        self.selectedRow = 0

        self.search = QLineEdit()
        self.search.textChanged.connect(self.findName)

        self.sheet_list = QComboBox()
        self.sheet_list.setPlaceholderText("Листы")
        self.sheet_list.currentTextChanged.connect(self.listChange)

        #Добавление виджетов pyqt6 в макет/каркас
        layout.addWidget(self.sheet_list)
        layout.addWidget(self.filter_list)
        layout.addWidget(self.search)
        layout.addWidget(self.addData)
        layout.addWidget(self.remove_row)
        
        path = "./data.xlsx"
        self.workbook = openpyxl.load_workbook(path)
        self.temp_sheets = {}

        # Предзагрузка листов
        for name in self.workbook.sheetnames:
            tableWidget = QTableWidget()

            tableWidget.cellClicked.connect(self.cellClicked)

            self.temp_sheets[self.workbook[name]] = tableWidget
            self.sheet_list.addItem(name)
            tableWidget.hide()

            layout.addWidget(tableWidget)

        self.currentWidget :QTableWidget = self.temp_sheets[next(iter(self.temp_sheets))]

        self.load_data()

    #Запуск окна с добавлением инфы
    def show_data_window(self):
        self.DataWindow.show()
        self.DataWindow.load()

    #Сортировка
    def findName(self):
        name = self.search.text().lower()
        found = False

        #Поиск по заданной строке
        for row in range(self.currentWidget.rowCount()):
            for column in range(self.currentWidget.columnCount()):
                header = self.currentWidget.horizontalHeaderItem(column)
                if header != None:
                    if header.text() == self.filter_list.currentText(): 
                        item = self.currentWidget.item(row, column)
                        try:
                            found = name in item.text().lower()
                            self.currentWidget.setRowHidden(row, not found)
                            if found:
                                break
                        except AttributeError:
                            pass

    #Переключение между листами Excel при помощи QListWidget
    def listChange(self, item):

        for sheet, widget in self.temp_sheets.items():
            widget.hide()

        self.temp_sheets[self.workbook[item]].show()
        self.currentWidget = self.temp_sheets[self.workbook[item]]

        labels = []

        for c in range(self.currentWidget.columnCount()):
            label = self.currentWidget.horizontalHeaderItem(c)
            labels.append(str(c+1) if label is None else label.text())
    
        #Загрузка фильтра при смене листа
        self.filter_list.clear()

        for i in labels:
            self.filter_list.addItem(i)

    #Загрузка всех листов Excel при помощи .loadSheet()
    def load_data(self):

        for name in self.workbook.sheetnames:
            self.loadSheet(self.workbook[name])

    #Загрузить лист Excel
    def loadSheet(self, sheet):

        self.temp_sheets[sheet].setRowCount(sheet.max_row)
        self.temp_sheets[sheet].setColumnCount(sheet.max_column)
        
        list_values = list(sheet.values)
        self.temp_sheets[sheet].setHorizontalHeaderLabels(list_values[0])
        self.temp_sheets[sheet].setSortingEnabled(True)

        row_index = 0

        # Добавление информации из листа Excel в QTableWidget
        for value_tuple in list_values[1:]:
            col_index = 0
            for value in value_tuple:
                item = TableWidgetItem(str(value))
                
                self.temp_sheets[sheet].setItem(row_index , col_index, item)
               
                col_index += 1
            row_index += 1

    #Добавить линию
    def addRow(self):
        self.currentWidget.insertRow(self.selectedRow)

        return self.selectedRow

    #Удалить линию
    def deleteRow(self, *args):
        
        self.acceptRemoveWindow.show()
        print(*args)
    
    # Сохранить текущую клетку
    def cellClicked(self, row, column):
        self.selectedRow = row
        

#Запуск кода
if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = Window()
    widget = Main()

    window.setCentralWidget(widget)
    widget.show()
    window.show()
    
    app.exec()