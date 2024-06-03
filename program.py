from PyQt6.QtCore import *
from PyQt6.QtWidgets import *
from PyQt6 import uic
import openpyxl
from datetime import date

class TableWidgetItem(QTableWidgetItem):
    def __lt__(self, other):
        try:
            return float(self.text()) < float(other.text())
        except ValueError:
            return super().__lt__(other)
        
class InfoWindow(QWidget):
    def __init__(self):
        super().__init__()

        self.TextLabel = QLabel()
        self._layount = QVBoxLayout()

        self._layount.addWidget(self.TextLabel)

        self.setLayout(self._layount)

    def ShowText(self, text):
        self.show()
        self.TextLabel.setText(text)


class AddClientWindow(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Add Client")

        self.Layout = QVBoxLayout()
        self.setLayout(self.Layout)

        self.addButton = QPushButton("Добавить")
        self.cancelButton = QPushButton("Отмена")
        self.cancelButton.clicked.connect(self.cancel)
        self.addButton.clicked.connect(self.acceptInfo)
        self.Layout.addWidget(self.addButton)
        self.Layout.addWidget(self.cancelButton)
        
        self.inputs :list[QLineEdit] = []

    def cancel(self):
        self.close()

    # Добавления новой строки по записанным данным
    def acceptInfo(self):
        sheet = sheets[0]

        rows = list(sheet.rows)

        for i in range(len(self.inputs)):
            value = self.inputs[i].text()

            sheet.cell(len(rows) + 1, i+1, value)
    
    #Загрузка полей для заполнений
    def load(self):
        print("loading stuff")

        self.Layout.removeWidget(self.addButton)
        self.Layout.removeWidget(self.cancelButton)

        labels = list(sheets[0].values)
        
        print(labels)

        for l in labels[0]:
            print(l)

            lineEdit = QLineEdit()
            lineEdit.setPlaceholderText(l)
            self.Layout.addWidget(lineEdit)

            self.inputs.append(lineEdit)

        self.Layout.addWidget(self.addButton)
        self.Layout.addWidget(self.cancelButton)

    # Действия при закрытии окна
    def closeEvent(self, event):
        print("clearing stuff")

        for i in reversed(range(self.Layout.count())): 
            self.Layout.itemAt(i).widget().setParent(None)

        self.inputs :list[QLineEdit] = []

        event.accept()

def loadClients():
    loadSheet(0)

def loadService():
    loadSheet(1)

def loadDone():
    loadSheet(2)

def programInfo():
    infoWindow.ShowText("ОБ ПРОГРАММЕ!!!")

def authorInfo():
    infoWindow.ShowText("ОБ АВТОРЕ!!!!!")

def export():
    filename, filter = QFileDialog.getSaveFileName(window, 'Save file', '','Excel files (*.xlsx)')
    
    workbook.save(filename)

def loadSheet(index):
    tableWidget.reset()

    sheet = sheets[index]
    global currentSheetName
    currentSheetName = workbook.sheetnames[index]

    global currentSheetIndex
    currentSheetIndex = index

    tableWidget.setRowCount(sheet.max_row - 1)
    tableWidget.setColumnCount(sheet.max_column)
    
    list_values = list(sheet.values)
    print(list_values[0])
    tableWidget.setHorizontalHeaderLabels(list_values[0])
    tableWidget.setSortingEnabled(True)

    filter_list.addItems(list_values[0])


  #  Добавление информации из листа Excel в QTableWidget

    row_index = 0
    for value_tuple in list_values[1:]:
        col_index = 0
        
        for value in value_tuple:
            item = TableWidgetItem(str(value))
            
            tableWidget.setItem(row_index , col_index, item)
            
            col_index += 1
        row_index += 1
    

def addClient():
    addClientWindow.load()
    addClientWindow.show()

def deleteRow():
    selectedLine = tableWidget.currentRow() + 1

    if selectedLine <= 0:
        return
    
    sheet = sheets[currentSheetIndex]
    sheet.delete_rows(selectedLine + 1)

    tableWidget.removeRow(selectedLine - 1)
    
    

def findName(self):
        name = search.text().lower()
        found = False

        #Поиск по заданной строке
        for row in range(tableWidget.rowCount()):
            for column in range(tableWidget.columnCount()):
                header = tableWidget.horizontalHeaderItem(column)
                if header != None:
                    if header.text() == filter_list.currentText():
                        item = tableWidget.item(row, column)
                        try:
                            found = name in item.text().lower()
                            tableWidget.setRowHidden(row, not found)
                            if found:
                                break
                        except AttributeError:
                            pass

Form, Window = uic.loadUiType("window.ui")
app = QApplication([])
window :QMainWindow = Window()
addClientWindow :QWidget = AddClientWindow()
infoWindow :InfoWindow = InfoWindow()

form = Form()
form.setupUi(window)

path = "./tabl.xlsx"
workbook = openpyxl.load_workbook(path)

sheets = workbook.worksheets

global currentSheetName  
currentSheetName = workbook.sheetnames[0]

global currentSheetIndex
currentSheetIndex = 0

tableWidget :QTableWidget = form.tableWidget

filter_list :QComboBox = form.comboBox_2

search :QLineEdit = form.lineEdit

comboBox :QComboBox = form.comboBox

for name in workbook.sheetnames:
    comboBox.addItem(name)

comboBox.currentIndexChanged.connect(loadSheet)

form.lineEdit.textChanged.connect(findName)

form.action_3.triggered.connect(export)
form.action.triggered.connect(authorInfo)
form.action_2.triggered.connect(programInfo)

form.pushButton.clicked.connect(addClient)
form.pushButton_2.clicked.connect(deleteRow)

window.show()
app.exec()